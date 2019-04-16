# -*-coding:utf-8 -*-

import os
import time
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from Config.ProjVar import *


class ParseExcel(object):
    """excel操作类"""
    def __init__(self,file_path):
        if not os.path.exists(file_path):
            self.wb = None
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.sheet = self.wb[self.wb.sheetnames[0]]

    def get_excel_file_path(self):
        """获取excel文件命名"""
        return self.file_path

    def set_sheet_by_name(self, name):
        """切换到name指定的sheet"""
        self.sheet = None
        if name in self.wb.sheetnames:
            self.sheet = self.wb[name]
        return self.sheet

    def get_current_sheet_name(self):
        """获取当前sheet的名称"""
        if self.sheet:
            return self.sheet.title

    def get_all_sheet_names(self):
        """获取所有的sheet名称"""
        return self.wb.sheetnames

    def get_min_row(self):
        """获取最小行号，从1开始"""
        try:
            return self.sheet.min_row
        except:
            traceback.print_exc()
            return None

    def get_max_row(self):
        """获取最大行号"""
        try:
            return self.sheet.max_row
        except:
            traceback.print_exc()
            return None

    def get_min_col(self):
        """获取最小列号，从1开始"""
        try:
            return self.sheet.min_column
        except:
            traceback.print_exc()
            return None

    def get_max_col(self):
        """获取最大列号"""
        try:
            return self.sheet.max_column
        except:
            traceback.print_exc()
            return None

    def get_row(self, row_no):
        """按索引获取指定的行"""
        if not isinstance(row_no, int):
            return None
        try:
            return list(self.sheet.rows)[row_no - 1]
        except:
            traceback.print_exc()

    def get_col(self, col_no):
        """按索引获取指定的列"""
        if not isinstance(col_no, int):
            return None
        try:
            return list(self.sheet.columns)[col_no - 1]
        except:
            traceback.print_exc()

    def get_cell_value(self, row_no, col_no):
        """获取指定单元格的内容"""
        if isinstance(row_no, int) and isinstance(col_no, int):
            try:
                return self.sheet.cell(row=row_no, column=col_no).value
            except:
                traceback.print_exc()

    def write_cell(self, row_no, col_no, value):
        """设置指定单元格的内容"""
        if isinstance(row_no, int) and isinstance(col_no, int):
            try:
                self.sheet.cell(row=row_no, column=col_no).value = value
                if "pass" == value.strip().lower():
                    self.set_cell_style(row_no=row_no, col_no=col_no, font=Font(color=colors.GREEN))
                elif "fail" == value.strip().lower():
                    self.set_cell_style(row_no=row_no, col_no=col_no, font=Font(color=colors.RED))
                self.save()
            except:
                traceback.print_exc()

    def write_cell_date(self, row_no, col_no):
        """向指定单元格写入日期字符串"""
        time_tuple = time.localtime()
        current_date = str(time_tuple.tm_year)+'年'+str(time_tuple.tm_mon)+'月'+ \
                      str(time_tuple.tm_mday) + '日'
        self.write_cell(row_no, col_no, current_date)

    def write_cell_time(self, row_no, col_no):
        """向指定单元格写入时间字符串"""
        time_tuple = time.localtime()
        current_time = str(time_tuple.tm_hour) + "时" + \
                      str(time_tuple.tm_min) + "分" + str(time_tuple.tm_sec) + "秒"
        self.write_cell(row_no, col_no, current_time)

    def write_cell_datetime(self, row_no, col_no):
        """向指定单元格写入日期和时间字符串"""
        time_tuple = time.localtime()
        current_date = str(time_tuple.tm_year) + "年" + \
                      str(time_tuple.tm_mon) + "月" + str(time_tuple.tm_mday) + "日"
        current_time = str(time_tuple.tm_hour) + "时" + \
                      str(time_tuple.tm_min) + "分" + str(time_tuple.tm_sec) + "秒"
        self.write_cell(row_no, col_no, current_date+" "+current_time)

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """合并单元格"""
        try:
            self.sheet.merge_cells(range_string, start_row, start_column, end_row, end_column)
        except:
            traceback.print_exc()
        else:
            self.save()

    def set_cell_style(self, row_no, col_no, border=None, fill=None,
                       font=None, alignment=None):
        """设定单元格的字体,颜色，边框，大小和边框背景色"""
        try:
            cell = self.sheet.cell(row=row_no, column=col_no)

            if border:
                cell.border = border
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
        except:
            traceback.print_exc()
        else:
            self.save()

    def save(self):
        """保存单元格"""
        self.wb.save(self.file_path)


if __name__ == "__main__":
    ws = ParseExcel(test_file_path)
    # print(ws.get_current_sheet_name())
    # print(ws.get_excel_file_path())
    # print(ws.get_min_row())
    ws.set_sheet_by_name("百度")
    print(ws.get_max_row())
    # print(ws.get_min_col())
    # print(ws.get_max_col())
    # print(ws.get_row(3))
    # print(ws.get_col(4))
    # ws.write_cell_date(-1, 100)
    # ws.write_cell_time(1, 2)
    # ws.write_cell_datetime(1, 3)
    # ws.merge_cells("A1:A3")
    # ws.set_cell_style(3, 3, font=Font(color=colors.RED))

